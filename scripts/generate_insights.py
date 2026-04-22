from __future__ import annotations

import argparse
import json
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any
from urllib.parse import quote

import requests
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = SCRIPT_DIR / "data"
MORPHOLOGY_DIR = DATA_DIR / "morphology"
WORDS_DIR = ROOT / "assets" / "words"
INSIGHTS_DIR = ROOT / "assets" / "insights"

WIKTAPI_BASE = "https://api.wiktapi.dev/v1/en/word"
DICTIONARY_API_BASE = "https://api.dictionaryapi.dev/api/v2/entries/en"
LEVELS = ("A1", "A2", "B1", "B2", "C1", "C2")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate offline word insights and morpheme hints.",
    )
    parser.add_argument(
        "--levels",
        nargs="*",
        default=list(LEVELS),
        help="CEFR levels to generate, e.g. A1 B1 C2",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=12,
        help="Parallel worker count for online definition fetches.",
    )
    parser.add_argument(
        "--skip-fetch",
        action="store_true",
        help="Generate morpheme data only and skip online definition fetches.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Optional max words per level for smoke testing.",
    )
    return parser.parse_args()


def load_words_by_level(levels: list[str]) -> dict[str, list[dict[str, Any]]]:
    result: dict[str, list[dict[str, Any]]] = {}
    for level in levels:
        path = WORDS_DIR / f"{level.lower()}_words.json"
        with path.open("r", encoding="utf-8") as handle:
            result[level] = json.load(handle)
    return result


def normalize_token(value: str) -> str:
    return re.sub(r"[^a-z-]", "", value.lower().strip())


def split_related_words(value: str) -> list[str]:
    tokens = [normalize_token(item) for item in value.split(";")]
    return [token for token in tokens if token]


def parse_morpheme_sheet(
    path: Path,
    *,
    kind: str,
    form_key: str,
    vocabulary: set[str],
) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [cell for cell in header_row]
    idx_form = headers.index(form_key)
    idx_meaning = headers.index("Meaning")
    idx_related = headers.index("Related Words")
    idx_theme = headers.index("Theme Category")

    result: dict[str, list[dict[str, Any]]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        form = str(row[idx_form] or "").strip()
        meaning = str(row[idx_meaning] or "").strip()
        theme = str(row[idx_theme] or "").strip()
        related_words = split_related_words(str(row[idx_related] or ""))
        if not form or not related_words:
            continue

        for matched_word in related_words:
            if matched_word not in vocabulary:
                continue
            siblings = [word for word in related_words if word != matched_word][:6]
            hint = {
                "kind": kind,
                "form": form,
                "meaning": meaning,
                "themeCategory": theme,
                "matchedWord": matched_word,
                "relatedWords": siblings,
            }
            result.setdefault(matched_word, []).append(hint)
    return result


def build_morpheme_index(vocabulary: set[str]) -> dict[str, list[dict[str, Any]]]:
    index: dict[str, list[dict[str, Any]]] = {}
    sources = [
        (
            MORPHOLOGY_DIR / "prefixes_rootionary_with_theme.xlsx",
            {"kind": "prefix", "form_key": "Prefix"},
        ),
        (
            MORPHOLOGY_DIR / "roots_rootionary_with_categories_en_final.xlsx",
            {"kind": "root", "form_key": "Root"},
        ),
        (
            MORPHOLOGY_DIR / "suffixes_rootionary_with_theme.xlsx",
            {"kind": "suffix", "form_key": "Suffix"},
        ),
    ]

    for path, options in sources:
        mapped = parse_morpheme_sheet(path, vocabulary=vocabulary, **options)
        for word, hints in mapped.items():
            index.setdefault(word, []).extend(hints)

    rank = {"root": 0, "prefix": 1, "suffix": 2}
    for hints in index.values():
        hints.sort(key=lambda hint: rank.get(hint["kind"], 9))
    return index


def fetch_json(session: requests.Session, url: str) -> Any:
    response = session.get(url, timeout=12)
    response.raise_for_status()
    return response.json()


def candidate_forms(raw: str) -> list[str]:
    word = raw.lower().strip()
    seen = {word}
    out = [word]

    def add(value: str) -> None:
        if len(value) >= 2 and value not in seen:
            seen.add(value)
            out.append(value)

    if word.endswith("ing") and len(word) > 5:
        stem = word[:-3]
        add(f"{stem}e")
        if len(stem) >= 2 and stem[-1] == stem[-2]:
            add(stem[:-1])
        add(stem)

    if word.endswith("ied") and len(word) > 4:
        add(f"{word[:-3]}y")
    elif word.endswith("ed") and len(word) > 4:
        stem = word[:-2]
        add(f"{stem}e")
        if len(stem) >= 2 and stem[-1] == stem[-2]:
            add(stem[:-1])
        add(stem)

    if word.endswith("ies") and len(word) > 4:
        add(f"{word[:-3]}y")
    elif word.endswith("ves") and len(word) > 4:
        add(f"{word[:-3]}f")
        add(f"{word[:-3]}fe")
    elif word.endswith("es") and len(word) > 3:
        add(word[:-2])
        add(word[:-1])
    elif word.endswith("s") and len(word) > 3 and not word.endswith("ss"):
        add(word[:-1])

    if word.endswith("ily") and len(word) > 5:
        add(f"{word[:-3]}y")
    elif word.endswith("ly") and len(word) > 4:
        add(word[:-2])

    if word.endswith("est") and len(word) > 5:
        add(word[:-3])
        add(f"{word[:-3]}e")
    elif word.endswith("er") and len(word) > 4 and not word.endswith("eer") and not word.endswith("ier"):
        add(word[:-2])
        add(f"{word[:-2]}e")

    if word.endswith("ation") and len(word) > 7:
        add(f"{word[:-5]}e")
        add(word[:-5])
    elif word.endswith("tion") and len(word) > 6:
        add(f"{word[:-4]}t")

    return out[:5]


def parse_wiktapi_entry(payload: Any) -> dict[str, str]:
    entries = payload.get("entries", []) if isinstance(payload, dict) else []
    for entry in entries:
        phonetic = ""
        sounds = entry.get("sounds", []) or []
        for sound in sounds:
            if sound.get("ipa"):
                phonetic = sound["ipa"]
                break

        part_of_speech = entry.get("pos", "") or ""
        senses = entry.get("senses", []) or []
        for sense in senses:
            glosses = sense.get("glosses", []) or []
            definition = next((gloss for gloss in glosses if gloss), "")
            examples = sense.get("examples", []) or []
            example = ""
            for candidate in examples:
                if isinstance(candidate, dict) and candidate.get("text"):
                    example = candidate["text"]
                    break
                if isinstance(candidate, str) and candidate:
                    example = candidate
                    break
            if definition:
                return {
                    "phonetic": phonetic,
                    "partOfSpeech": part_of_speech,
                    "definition": definition,
                    "example": example,
                    "source": "wiktapi",
                }
    return {}


def parse_dictionary_api_entry(payload: Any) -> dict[str, str]:
    if not isinstance(payload, list):
        return {}

    for entry in payload:
        phonetic = entry.get("phonetic", "") or ""
        if not phonetic:
            for item in entry.get("phonetics", []) or []:
                phonetic = item.get("text", "") or ""
                if phonetic:
                    break

        for meaning in entry.get("meanings", []) or []:
            definitions = meaning.get("definitions", []) or []
            for definition_entry in definitions:
                definition = definition_entry.get("definition", "") or ""
                example = definition_entry.get("example", "") or ""
                if definition:
                    return {
                        "phonetic": phonetic,
                        "partOfSpeech": meaning.get("partOfSpeech", "") or "",
                        "definition": definition,
                        "example": example,
                        "source": "dictionaryapi",
                    }
    return {}


def fetch_definition(session: requests.Session, word: str) -> dict[str, str]:
    for candidate in candidate_forms(word):
        try:
            payload = fetch_json(session, f"{DICTIONARY_API_BASE}/{quote(candidate)}")
            parsed = parse_dictionary_api_entry(payload)
            if parsed:
                if candidate != word:
                    parsed["source"] = f'dictionaryapi:{candidate}'
                return parsed
        except requests.RequestException:
            pass

    for candidate in candidate_forms(word):
        try:
            payload = fetch_json(session, f"{WIKTAPI_BASE}/{quote(candidate)}")
            parsed = parse_wiktapi_entry(payload)
            if parsed:
                if candidate != word:
                    parsed["source"] = f'wiktapi:{candidate}'
                return parsed
        except requests.RequestException:
            pass

    return {
        "phonetic": "",
        "partOfSpeech": "",
        "definition": "",
        "example": "",
        "source": "",
    }


def build_insight(
    session: requests.Session | None,
    word_record: dict[str, Any],
    morpheme_index: dict[str, list[dict[str, Any]]],
) -> dict[str, Any]:
    word = str(word_record["word"]).lower()
    fetched = (
        fetch_definition(session, word) if session is not None else {
            "phonetic": "",
            "partOfSpeech": "",
            "definition": "",
            "example": "",
            "source": "",
        }
    )
    morphemes = morpheme_index.get(word, [])

    return {
        "wordId": word_record["id"],
        "phonetic": fetched["phonetic"],
        "partOfSpeech": fetched["partOfSpeech"],
        "definition": fetched["definition"],
        "example": fetched["example"],
        "source": fetched["source"],
        "hasInsight": bool(fetched["definition"] or morphemes),
        "morphemes": morphemes,
    }


def generate_level(
    *,
    level: str,
    words: list[dict[str, Any]],
    morpheme_index: dict[str, list[dict[str, Any]]],
    workers: int,
    skip_fetch: bool,
) -> list[dict[str, Any]]:
    session = None if skip_fetch else requests.Session()
    output: list[dict[str, Any]] = [None] * len(words)  # type: ignore[list-item]

    if skip_fetch:
        for index, word_record in enumerate(words):
            output[index] = build_insight(None, word_record, morpheme_index)
        return output

    with session:
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {
                executor.submit(build_insight, session, word_record, morpheme_index): index
                for index, word_record in enumerate(words)
            }
            completed = 0
            started_at = time.time()
            for future in as_completed(futures):
                index = futures[future]
                output[index] = future.result()
                completed += 1
                if completed % 100 == 0 or completed == len(words):
                    elapsed = max(time.time() - started_at, 1)
                    rate = completed / elapsed
                    print(
                        f"[{level}] {completed}/{len(words)} completed "
                        f"({rate:.1f} words/sec)",
                        flush=True,
                    )
    return output


def write_level(level: str, insights: list[dict[str, Any]]) -> None:
    INSIGHTS_DIR.mkdir(parents=True, exist_ok=True)
    path = INSIGHTS_DIR / f"{level.lower()}_insights.json"
    tmp_path = path.with_suffix(".json.tmp")
    with tmp_path.open("w", encoding="utf-8") as handle:
        json.dump(insights, handle, ensure_ascii=False, indent=2)
    tmp_path.replace(path)


def main() -> None:
    args = parse_args()
    requested_levels = [level.upper() for level in args.levels]
    invalid = [level for level in requested_levels if level not in LEVELS]
    if invalid:
        raise SystemExit(f"Unsupported levels: {', '.join(invalid)}")

    words_by_level = load_words_by_level(requested_levels)
    if args.limit > 0:
        words_by_level = {
            level: records[: args.limit] for level, records in words_by_level.items()
        }

    vocabulary = {
        str(record["word"]).lower()
        for records in words_by_level.values()
        for record in records
    }
    morpheme_index = build_morpheme_index(vocabulary)

    for level in requested_levels:
        print(
            f"Generating {level} insights for {len(words_by_level[level])} words...",
            flush=True,
        )
        insights = generate_level(
            level=level,
            words=words_by_level[level],
            morpheme_index=morpheme_index,
            workers=max(args.workers, 1),
            skip_fetch=args.skip_fetch,
        )
        write_level(level, insights)
        filled = sum(1 for item in insights if item["definition"])
        print(
            f"[{level}] wrote {filled}/{len(insights)} definitions to assets.",
            flush=True,
        )

    matched_words = sum(1 for hints in morpheme_index.values() if hints)
    print(f"Morpheme hints matched {matched_words} words.", flush=True)
    print("Done.", flush=True)


if __name__ == "__main__":
    main()
